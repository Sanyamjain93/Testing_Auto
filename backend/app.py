"""
FastAPI backend — AI Test Case Generator with session-based authentication.

Each user gets an isolated session: own input/output dirs, own pipeline state,
own LLM selection, own SSE queues. Data is in-memory only, destroyed on logout.
"""

import contextlib
import io
import json
import logging
import mimetypes
import queue
import re
import shutil
import sys
import threading
import uuid
import zipfile
from pathlib import Path
from typing import AsyncGenerator

import pandas as pd
import uvicorn
from fastapi import FastAPI, File, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT           = Path(__file__).parent.parent
SESSIONS_BASE  = ROOT / "data" / "sessions"
USERS_FILE     = ROOT / "users.xlsx"
FRONTEND_BUILD = ROOT / "frontend" / "dist"

sys.path.insert(0, str(ROOT))

from logger import get_logger, new_run_log  # noqa: E402
from util.mistral_client import LLM_OPTIONS, MistralLLM  # noqa: E402
from util.playwright_generator import generate_grouped_script, generate_playwright_script  # noqa: E402

logger = get_logger("test_automation.backend")

# ── FastAPI app ───────────────────────────────────────────────────────────────
app = FastAPI(title="AI Test Case Generator API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://test-automation-2251.azurewebsites.net",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_ALLOWED_EXT = {".pdf", ".docx", ".txt", ".md"}

# ── User authentication ────────────────────────────────────────────────────────
def _load_users() -> pd.DataFrame:
    if USERS_FILE.exists():
        return pd.read_excel(USERS_FILE, dtype=str)
    logger.warning("users.xlsx not found — no users can log in")
    return pd.DataFrame(columns=["username", "password"])


_users_df: pd.DataFrame = _load_users()


def validate_user(username: str, password: str) -> bool:
    mask = (_users_df["username"] == username) & (_users_df["password"] == password)
    return bool(mask.any())


# ── Session store ──────────────────────────────────────────────────────────────
user_sessions: dict = {}


def _session_dir(sid: str) -> Path:
    return SESSIONS_BASE / sid


def _create_session(username: str) -> tuple[str, dict]:
    sid = str(uuid.uuid4())
    sess_dir = _session_dir(sid)
    return sid, {
        "username": username,
        "state":    {"running": False, "done": False, "error": None},
        "lock":     threading.Lock(),
        "log_q":    queue.Queue(),
        "progress_q": queue.Queue(),
        "selected_llm": LLM_OPTIONS[3].copy(),
        "input_dir":         sess_dir / "input",
        "output_file":       sess_dir / "output.xlsx",
        "scripts_dir":       sess_dir / "scripts",
        "faiss_index_file":  sess_dir / "faiss_index.faiss",
        "rag_metadata_file": sess_dir / "metadata.json",
    }


def get_current_session(request: Request) -> tuple[str, dict]:
    """Validate session cookie. Raises HTTP 401 if missing or unknown."""
    sid = request.cookies.get("session_id")
    if not sid or sid not in user_sessions:
        raise HTTPException(status_code=401, detail="Not authenticated. Please log in.")
    return sid, user_sessions[sid]


# ── Pydantic models ────────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    username: str
    password: str


class LLMSelection(BaseModel):
    provider: str
    model: str


# ── Pipeline infrastructure ────────────────────────────────────────────────────
def _make_emit(progress_q: queue.Queue):
    def emit(stage: str, status: str) -> None:
        progress_q.put({"stage": stage, "status": status})
    return emit


class _QueueWriter(io.TextIOBase):
    """Redirect print() into a per-session SSE log queue."""

    def __init__(self, log_q: queue.Queue) -> None:
        self._q = log_q

    def write(self, s: str) -> int:
        if s.strip():
            self._q.put(s.rstrip())
        return len(s)


def _pipeline_thread(sid: str) -> None:
    if sid not in user_sessions:
        return
    sess   = user_sessions[sid]
    writer = _QueueWriter(sess["log_q"])
    with sess["lock"]:
        provider = sess["selected_llm"]["provider"]
        model    = sess["selected_llm"]["model"]
    log_file = new_run_log()
    logger.info(f"[PIPELINE][{sess['username']}] Run log: {log_file}")
    try:
        sess["input_dir"].mkdir(parents=True, exist_ok=True)
        with contextlib.redirect_stdout(writer):
            from pipeline import run  # noqa: PLC0415
            run(
                provider=provider,
                model=model,
                emit_progress=_make_emit(sess["progress_q"]),
                input_dir=str(sess["input_dir"]),
                output_file=str(sess["output_file"]),
                faiss_index_file=str(sess["faiss_index_file"]),
                rag_metadata_file=str(sess["rag_metadata_file"]),
            )
        with sess["lock"]:
            sess["state"].update({"running": False, "done": True, "error": None})
    except Exception as exc:  # noqa: BLE001
        logger.error("Pipeline failed", exc_info=True)
        sess["log_q"].put(f"❌ Pipeline error: {exc}")
        with sess["lock"]:
            sess["state"].update({"running": False, "done": True, "error": str(exc)})


# ── Auth endpoints ─────────────────────────────────────────────────────────────────────────────

@app.post("/login")
def login(body: LoginRequest, response: Response) -> dict:
    if not validate_user(body.username, body.password):
        raise HTTPException(status_code=401, detail="Invalid username or password.")
    sid, sess = _create_session(body.username)
    user_sessions[sid] = sess
    response.set_cookie(
        key="session_id",
        value=sid,
        httponly=True,
        samesite="lax",
        secure=False,   # set True in production behind HTTPS
        max_age=86400,  # 24-hour expiry
    )
    logger.info(f"[AUTH] '{body.username}' logged in (sid={sid[:8]}…)")
    return {"message": "Login successful", "username": body.username}


@app.post("/logout")
def logout(request: Request, response: Response) -> dict:
    sid = request.cookies.get("session_id")
    if sid and sid in user_sessions:
        username = user_sessions[sid]["username"]
        sess_dir = _session_dir(sid)
        user_sessions.pop(sid, None)
        if sess_dir.exists():
            shutil.rmtree(sess_dir, ignore_errors=True)
        logger.info(f"[AUTH] '{username}' logged out (sid={sid[:8]}…)")
    response.delete_cookie("session_id")
    return {"message": "Logged out successfully"}


@app.get("/me")
def get_me(request: Request) -> dict:
    _, sess = get_current_session(request)
    return {"username": sess["username"], "authenticated": True}


# ── Pipeline endpoints ─────────────────────────────────────────────────────────────────────────

@app.get("/status")
def get_status(request: Request) -> dict:
    _, sess = get_current_session(request)
    return sess["state"]


@app.get("/llm-options")
def get_llm_options(request: Request) -> dict:
    _, sess = get_current_session(request)
    return {"options": LLM_OPTIONS, "selected": sess["selected_llm"]}


@app.post("/set-llm")
def set_llm(selection: LLMSelection, request: Request) -> dict:
    _, sess = get_current_session(request)
    valid = any(
        o["provider"] == selection.provider and o["model"] == selection.model
        for o in LLM_OPTIONS
    )
    if not valid:
        raise HTTPException(status_code=400, detail="Unknown provider/model combination.")
    with sess["lock"]:
        if sess["state"]["running"]:
            raise HTTPException(status_code=409, detail="Pipeline is running. Cannot change LLM now.")
        sess["selected_llm"].update({"provider": selection.provider, "model": selection.model})
    logger.info(f"[LLM][{sess['username']}] Selected: {selection.provider} / {selection.model}")
    return {"status": "ok", "selected": sess["selected_llm"]}


@app.post("/upload")
async def upload_files(request: Request, files: list[UploadFile] = File(...)) -> dict:
    _, sess = get_current_session(request)
    session_input_dir = sess["input_dir"]
    session_input_dir.mkdir(parents=True, exist_ok=True)

    for f in files:
        ext = Path(f.filename).suffix.lower()
        if ext not in _ALLOWED_EXT:
            raise HTTPException(
                status_code=400,
                detail=f"'{ext}' is not allowed. Accepted: {sorted(_ALLOWED_EXT)}",
            )

    staging = session_input_dir.parent / "_staging"
    if staging.exists():
        shutil.rmtree(staging)
    staging.mkdir(parents=True)

    saved: list[str] = []
    try:
        for f in files:
            dest = staging / Path(f.filename).name
            with dest.open("wb") as out:
                shutil.copyfileobj(f.file, out)
            saved.append(f.filename)
    except Exception as exc:
        shutil.rmtree(staging, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"File save failed: {exc}") from exc

    for existing in session_input_dir.iterdir():
        existing.unlink()
    for staged_file in staging.iterdir():
        shutil.move(str(staged_file), session_input_dir / staged_file.name)
    shutil.rmtree(staging, ignore_errors=True)

    logger.info(f"[UPLOAD][{sess['username']}] {len(saved)} file(s): {saved}")
    return {"uploaded": saved}


@app.post("/reset")
def reset_state(request: Request) -> dict:
    """Reset per-session pipeline state and clear queues."""
    _, sess = get_current_session(request)
    with sess["lock"]:
        if sess["state"]["running"]:
            raise HTTPException(
                status_code=409,
                detail="Pipeline is running. Wait for it to finish before resetting.",
            )
        sess["state"].update({"running": False, "done": False, "error": None})
    while not sess["log_q"].empty():
        sess["log_q"].get_nowait()
    while not sess["progress_q"].empty():
        sess["progress_q"].get_nowait()
    logger.info(f"[RESET][{sess['username']}] State cleared.")
    return {"status": "reset"}


@app.post("/run")
def start_pipeline(request: Request) -> dict:
    sid, sess = get_current_session(request)
    with sess["lock"]:
        if sess["state"]["running"]:
            return {"status": "already_running"}
        sess["state"].update({"running": True, "done": False, "error": None})

    while not sess["log_q"].empty():
        sess["log_q"].get_nowait()
    while not sess["progress_q"].empty():
        sess["progress_q"].get_nowait()

    thread = threading.Thread(target=_pipeline_thread, args=(sid,), daemon=True)
    thread.start()
    return {"status": "started"}


@app.get("/stream")
async def stream_logs(request: Request) -> StreamingResponse:
    import asyncio  # noqa: PLC0415

    _, sess     = get_current_session(request)
    log_q      = sess["log_q"]
    progress_q = sess["progress_q"]
    state      = sess["state"]

    async def _generate() -> AsyncGenerator[str, None]:
        while True:
            try:
                prog = progress_q.get_nowait()
                yield f"data: {json.dumps({'progress': prog})}\n\n"
            except queue.Empty:
                pass

            try:
                msg = log_q.get_nowait()
                yield f"data: {json.dumps({'log': msg})}\n\n"
            except queue.Empty:
                if not state["running"]:
                    draining = True
                    while draining:
                        draining = False
                        try:
                            prog = progress_q.get_nowait()
                            yield f"data: {json.dumps({'progress': prog})}\n\n"
                            draining = True
                        except queue.Empty:
                            pass
                        try:
                            leftover = log_q.get_nowait()
                            yield f"data: {json.dumps({'log': leftover})}\n\n"
                            draining = True
                        except queue.Empty:
                            pass
                    break
                await asyncio.sleep(0.15)

        if state.get("error"):
            yield f"data: {json.dumps({'error': state['error']})}\n\n"
        yield f"data: {json.dumps({'done': True})}\n\n"

    return StreamingResponse(
        _generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.delete("/rag-index")
def refresh_rag_index(request: Request) -> dict:
    """Delete the per-session FAISS index so it is rebuilt on the next run."""
    _, sess = get_current_session(request)
    if sess["state"]["running"]:
        raise HTTPException(
            status_code=409,
            detail="Pipeline is currently running. Wait for it to finish.",
        )
    deleted = []
    for f in (sess["faiss_index_file"], sess["rag_metadata_file"]):
        if f.exists():
            f.unlink()
            deleted.append(f.name)
    logger.info(f"[RAG][{sess['username']}] Index cleared: {deleted or 'nothing to delete'}")
    return {"deleted": deleted, "message": "RAG index cleared. It will be rebuilt on the next Run."}


@app.post("/generate-scripts")
def generate_scripts_endpoint(request: Request) -> dict:
    """Read test cases from the per-session Excel, generate Playwright scripts."""
    _, sess = get_current_session(request)
    output_file  = sess["output_file"]
    scripts_dir  = sess["scripts_dir"]
    selected_llm = sess["selected_llm"]

    if not output_file.exists():
        raise HTTPException(
            status_code=404,
            detail="No test cases found. Run the pipeline first.",
        )

    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    tests_map: dict = {}
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, row_vals))
        key = (row.get("Requirement ID"), row.get("Test Name"), row.get("Test Description"))
        if key not in tests_map:
            tests_map[key] = {
                "requirement_id": row.get("Requirement ID", ""),
                "test_name": row.get("Test Name", ""),
                "test_description": row.get("Test Description", "") or "",
                "test_type": str(row.get("Test Type") or "positive").strip().lower(),
                "steps": [],
            }
        tests_map[key]["steps"].append({
            "step_name": row.get("Step Name", ""),
            "action": row.get("Action", ""),
            "expected_result": row.get("Expected Result", ""),
        })

    tests = list(tests_map.values())
    if not tests:
        raise HTTPException(status_code=404, detail="No test cases found in output file.")

    logger.info(f"[SCRIPT][{sess['username']}] Generating scripts for {len(tests)} test case(s)")
    llm = MistralLLM(selected_llm["provider"], selected_llm["model"])
    scripts_dir.mkdir(parents=True, exist_ok=True)

    groups: dict[str, list] = {}
    for t in tests:
        req_id = t.get("requirement_id") or "UNGROUPED"
        groups.setdefault(req_id, []).append(t)

    count  = 0
    errors: list[str] = []
    for group_idx, (req_id, group_tests) in enumerate(groups.items()):
        try:
            safe_req    = re.sub(r"[^\w\s-]", "", req_id).strip().replace(" ", "_")[:40]
            first_name  = group_tests[0]["test_name"]
            group_label = re.sub(r"^[\d_]+", "", first_name).strip().replace("_", " ")
            group_label = re.sub(r"\s+", " ", group_label)[:60].strip() + " Tests"
            if len(group_tests) == 1:
                script = generate_playwright_script(group_tests[0], llm)
            else:
                script = generate_grouped_script(group_label, group_tests, llm)
            filename = f"{group_idx + 1:03d}_{safe_req}.spec.js"
            (scripts_dir / filename).write_text(script, encoding="utf-8")
            count += 1
            logger.info(f"[SCRIPT][{sess['username']}] Generated: {filename}")
        except Exception as exc:  # noqa: BLE001
            failed_names = [t["test_name"] for t in group_tests]
            errors.extend(failed_names)
            logger.error(f"[SCRIPT ERROR][{sess['username']}] group={req_id}: {exc}", exc_info=True)

    return {"message": f"{count} script(s) generated", "count": count, "errors": errors}


@app.get("/download-scripts")
def download_scripts_endpoint(request: Request) -> FileResponse:
    """Zip all per-session Playwright scripts and return the archive."""
    _, sess = get_current_session(request)
    scripts_dir = sess["scripts_dir"]
    scripts = sorted(scripts_dir.glob("*.spec.js")) if scripts_dir.exists() else []
    if not scripts:
        raise HTTPException(
            status_code=404,
            detail="No scripts found. Click 'Generate Playwright Scripts' first.",
        )

    zip_path = sess["output_file"].parent / "playwright_scripts.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in scripts:
            zf.write(f, f.name)

    return FileResponse(str(zip_path), media_type="application/zip", filename="playwright_scripts.zip")


def _derive_script_label(script_name: str, script_content: str) -> str:
    """Build a user-friendly label from the first test title in the script."""
    match = re.search(r"test\s*\(\s*['\"]([^'\"]+)['\"]", script_content)
    if match:
        return match.group(1).strip()
    return Path(script_name).stem.replace("_", " ")


@app.get("/scripts")
def list_scripts(request: Request) -> dict:
    """Return per-session Playwright scripts for the in-app code viewer."""
    _, sess = get_current_session(request)
    scripts_dir = sess["scripts_dir"]
    scripts = sorted(scripts_dir.glob("*.spec.js")) if scripts_dir.exists() else []
    if not scripts:
        return {"scripts": []}

    items = []
    for script_path in scripts:
        content = script_path.read_text(encoding="utf-8")
        items.append({
            "name":     script_path.name,
            "label":    _derive_script_label(script_path.name, content),
            "language": "javascript",
            "content":  content,
            "size":     script_path.stat().st_size,
            "mime":     mimetypes.guess_type(script_path.name)[0] or "text/plain",
        })
    return {"scripts": items}


@app.get("/results")
def get_results(request: Request) -> dict:
    _, sess = get_current_session(request)
    output_file = sess["output_file"]
    if not output_file.exists():
        return {"test_cases": []}
    try:
        import openpyxl  # noqa: PLC0415
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        rows = [
            dict(zip(headers, row))
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        return {"test_cases": rows}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/download")
def download_excel(request: Request) -> FileResponse:
    _, sess = get_current_session(request)
    output_file = sess["output_file"]
    if not output_file.exists():
        raise HTTPException(
            status_code=404,
            detail="No output file found. Run the pipeline first.",
        )
    return FileResponse(
        str(output_file),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="generated_tests.xlsx",
    )


# ── Serve React SPA in production ─────────────────────────────────────────────
if FRONTEND_BUILD.exists():
    _assets = FRONTEND_BUILD / "assets"
    if _assets.exists():
        app.mount("/assets", StaticFiles(directory=str(_assets)), name="static-assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    async def spa_fallback(full_path: str) -> FileResponse:  # noqa: ARG001
        return FileResponse(str(FRONTEND_BUILD / "index.html"))
else:
    @app.get("/", include_in_schema=False)
    async def root() -> dict:
        return {"message": "Backend running. Start the frontend with: cd frontend && npm run dev"}


# ── entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True, app_dir=str(Path(__file__).parent))
